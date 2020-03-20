from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import re
def findInList(listToSearch, item):
    try:
        retval=listToSearch.index(item)
        #print(retval)
        return retval
    except:
        return -1

def getQuestionsAndOptions(questionsList):

    #nextQuestionStarted=False
    #hasReferenceChanged=False

    questionObject=[]
    questionDict={}
    options=0
    questionRef=""

    for question in questionsList:
        classList=question.get("class")
        #print(classList)
        if(findInList(classList,"quslist")>-1):        
            #nextQuestionStarted=True        
            questionObject.append(questionDict)
            questionDict={}
            
            questionReference=question.find("div",{"class": "qus_ref"})
            
            if(questionReference):
                #hasReferenceChanged=True
                questionRef=questionReference.text
            
            
            questionDict["questionReference"]=questionRef
            questionDict["question"]=question.text
            option=0
            
        

        if(findInList(classList,"optdiv")>-1):        
        
            questionDict["option_"+str(option)]=question.text
            option=option+1
            
        if(findInList(classList,"exp_text")>-1):   
            questionDict["explaination"]=question.text
            img=question.find("img")
            if(img):
                questionDict["image_explaination"]=img.get("src")

            
        if(findInList(classList,"crct")>-1):    
            questionDict["correct_answer"]=question.find("span").text


    questionObject.append(questionDict)
    
    return questionObject
def removeNonAscii(s):
    return "".join(i for i in s if ord(i)<128)


def cleanup(noisyText):
    cleanedString=noisyText
    cleanedString=cleanedString.replace("\n", "")
    cleanedString=removeNonAscii(cleanedString)
    #print(cleanedString)
    return cleanedString
    

URL_List=["https://www.fresherslive.com/online-test/alphabet-sequence-questions-and-answers/", "https://www.fresherslive.com/online-test/direction-sense-questions-and-answers/",'https://www.fresherslive.com/online-test/symbols-questions-and-answers/', 'https://www.fresherslive.com/online-test/series-questions-and-answers/']    
#pagesCount=[1,1,6,12]
pagesCount=[1,1,1,1]
workbook = Workbook()
rowNumber=0
for i in range(0, len(URL_List)):
    URL=URL_List[i]
    pages=pagesCount[i]
   

    
    sheet = workbook.active
    
    tempURL=URL
    tempURL=tempURL.replace("questions-and-answers", "")
    tempURL=tempURL.replace("-", " ")
    sheetName=tempURL.split("/")

    if(i==0):
        sheet.title=sheetName[-2].title()
        activeSheet=workbook.active
    else:
        workbook.create_sheet(sheetName[-2].title())
        workbook.active=i
        activeSheet=workbook.active
    rowNumber=1
    activeSheet.cell(row=rowNumber, column=1).value = "Question"
    activeSheet.cell(row=rowNumber, column=2).value = "Option 1"
    activeSheet.cell(row=rowNumber, column=3).value = "Option 2"
    activeSheet.cell(row=rowNumber, column=4).value = "Option 3"
    activeSheet.cell(row=rowNumber, column=5).value = "Option 4"
    activeSheet.cell(row=rowNumber, column=6).value = "Explaination"
    activeSheet.cell(row=rowNumber, column=7).value = "Question Reference"
    
    #activeSheet.cell(row=rowNumber, column=6).value = "Option 5"
    rowNumber=2
    for pagenumber in range (1,pages+1) :
        indexURL=URL+str(pagenumber)    
        html_doc=requests.get(indexURL)
        soup = BeautifulSoup(html_doc.text, 'html.parser')
        questionList=soup.find("div", {"class": "quswrap"}).find_all("div")
        extractedData=getQuestionsAndOptions(questionList)
        for data in extractedData:
            if(len(data)==0):
                continue
            activeSheet.cell(row=rowNumber, column=1).value = cleanup(data.get("question"))
            activeSheet.cell(row=rowNumber, column=2).value = cleanup(data.get("option_0"))
            activeSheet.cell(row=rowNumber, column=3).value = cleanup(data.get("option_1"))
            activeSheet.cell(row=rowNumber, column=4).value = cleanup(data.get("option_2"))
            activeSheet.cell(row=rowNumber, column=5).value = cleanup(data.get("option_3"))
            activeSheet.cell(row=rowNumber, column=6).value = cleanup(data.get("explaination"))
            activeSheet.cell(row=rowNumber, column=7).value = cleanup(data.get("questionReference"))
            
            rowNumber=rowNumber+1
        #print(extractedData)
    

workbook.save(filename="hello_world.xlsx")

